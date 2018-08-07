import openpyxl


class W(object):
    def __init__(self):
        self.wb = openpyxl.load_workbook('02-5月洗衣机行业占比累计数据_20180528.xlsx')
        # sheet_name = wb.get_sheet_names()[1]
        sheet_name = self.wb.sheetnames[1]
        # ws = wb.get_sheet_by_name(sheet_name)
        self.ws = self.wb[sheet_name]
        print(self.ws)
        print(self.ws['C1'].value)
        print(self.ws.max_row)
        print(self.ws.max_column)
        self.maxRow = self.ws.max_row


    def w(self, data):
        data_keys = list(data.keys())
        if '标识' in data and '每日' == data['标识']:
            if 'Haier/海尔' == data['品牌']:
                for i in range(1, 7):
                    self.ws.cell(row=self.maxRow + 2, column=i, value=data[data_keys[i]])
            elif 'Littleswan/小天鹅' == data['品牌']:
                for i in range(1, 7):
                    self.ws.cell(row=self.maxRow + 5, column=i, value=data[data_keys[i]])
            elif 'Midea/美的' == data['品牌']:
                for i in range(1, 7):
                    self.ws.cell(row=self.maxRow + 6, column=i, value=data[data_keys[i]])
            elif 'Sanyo/三洋' == data['品牌']:
                for i in range(1, 7):
                    self.ws.cell(row=self.maxRow + 7, column=i, value=data[data_keys[i]])
            elif 'TCL' == data['品牌']:
                for i in range(1, 7):
                    self.ws.cell(row=self.maxRow + 8, column=i, value=data[data_keys[i]])
            elif 'Panasonic/松下' == data['品牌']:
                for i in range(1, 7):
                    self.ws.cell(row=self.maxRow + 9, column=i, value=data[data_keys[i]])
            elif 'SIEMENS/西门子' == data['品牌']:
                for i in range(1, 7):
                    self.ws.cell(row=self.maxRow + 10, column=i, value=data[data_keys[i]])
            elif 'Leader/统帅' == data['品牌']:
                for i in range(1, 7):
                    self.ws.cell(row=self.maxRow + 11, column=i, value=data[data_keys[i]])
            elif 'Royalstar/荣事达' == data['品牌']:
                for i in range(1, 7):
                    self.ws.cell(row=self.maxRow + 12, column=i, value=data[data_keys[i]])
            elif 'Whirlpool/惠而浦' == data['品牌']:
                for i in range(1, 7):
                    self.ws.cell(row=self.maxRow + 13, column=i, value=data[data_keys[i]])
        elif '标识' in data and '累计' == data['标识']:
            if 'Haier/海尔' == data['品牌']:
                for i in range(7, 11):
                    self.ws.cell(row=self.maxRow + 2, column=i, value=data[data_keys[i - 4]])
            elif 'Littleswan/小天鹅' == data['品牌']:
                for i in range(7, 11):
                    self.ws.cell(row=self.maxRow + 5, column=i, value=data[data_keys[i - 4]])
            elif 'Midea/美的' == data['品牌']:
                for i in range(7, 11):
                    self.ws.cell(row=self.maxRow + 6, column=i, value=data[data_keys[i - 4]])
            elif 'Sanyo/三洋' == data['品牌']:
                for i in range(7, 11):
                    self.ws.cell(row=self.maxRow + 7, column=i, value=data[data_keys[i - 4]])
            elif 'TCL' == data['品牌']:
                for i in range(7, 11):
                    self.ws.cell(row=self.maxRow + 8, column=i, value=data[data_keys[i - 4]])
            elif 'Panasonic/松下' == data['品牌']:
                for i in range(7, 11):
                    self.ws.cell(row=self.maxRow + 9, column=i, value=data[data_keys[i - 4]])
            elif 'SIEMENS/西门子' == data['品牌']:
                for i in range(7, 11):
                    self.ws.cell(row=self.maxRow + 10, column=i, value=data[data_keys[i - 4]])
            elif 'Leader/统帅' == data['品牌']:
                for i in range(7, 11):
                    self.ws.cell(row=self.maxRow + 11, column=i, value=data[data_keys[i - 4]])
            elif 'Royalstar/荣事达' == data['品牌']:
                for i in range(7, 11):
                    self.ws.cell(row=self.maxRow + 12, column=i, value=data[data_keys[i - 4]])
            elif 'Whirlpool/惠而浦' == data['品牌']:
                for i in range(7, 11):
                    self.ws.cell(row=self.maxRow + 13, column=i, value=data[data_keys[i - 4]])


        elif '美的系' == data['品牌']:
            self.ws.cell(row=self.maxRow + 3, column=1, value=data['日期'])
            self.ws.cell(row=self.maxRow + 3, column=2, value=data['品牌'])
        elif '惠而浦系' == data['品牌']:
            self.ws.cell(row=self.maxRow + 4, column=1, value=data['日期'])
            self.ws.cell(row=self.maxRow + 4, column=2, value=data['品牌'])
        self.wb.save('1.xlsx')

    # def __del__(self):
    #     self.wb.save('1.xlsx')

data_list = [
    {'标识': '每日', '日期': '5月27日', '品牌': 'Haier/海尔', '客单价': '1646.79', '支付转化率': '1.73%', '访客数': '102075', '销售额': '2908062.344'},
    {'标识': '累计', '日期': '5月27日', '品牌': 'Haier/海尔', '客单价': '1782.24', '支付转化率': '2.81%', '访客数': '2709575', '销售额': '135698073.8'},
    {'标识': '每日', '日期': '5月27日', '品牌': 'Littleswan/小天鹅', '客单价': '1619.52', '支付转化率': '1.43%', '访客数': '78928', '销售额': '1827904.286'},
    {'标识': '累计', '日期': '5月27日', '品牌': 'Littleswan/小天鹅', '客单价': '1882.24', '支付转化率': '2.89%', '访客数': '2710575', '销售额': '135778073.8'},
    {'标识': '每日', '日期': '5月27日', '品牌': 'Midea/美的', '客单价': '1619.52', '支付转化率': '1.43%', '访客数': '78928', '销售额': '1827904.286'},
    {'标识': '累计', '日期': '5月27日', '品牌': 'Midea/美的', '客单价': '1882.24', '支付转化率': '2.89%', '访客数': '2710575', '销售额': '135778073.8'},
    {'标识': '每日', '日期': '5月27日', '品牌': 'Sanyo/三洋', '客单价': '1619.52', '支付转化率': '1.43%', '访客数': '78928', '销售额': '1827904.286'},
    {'标识': '累计', '日期': '5月27日', '品牌': 'Sanyo/三洋', '客单价': '1882.24', '支付转化率': '2.89%', '访客数': '2710575', '销售额': '135778073.8'},
    {'标识': '每日', '日期': '5月27日', '品牌': 'TCL', '客单价': '1619.52', '支付转化率': '1.43%', '访客数': '78928', '销售额': '1827904.286'},
    {'标识': '累计', '日期': '5月27日', '品牌': 'TCL', '客单价': '1882.24', '支付转化率': '2.89%', '访客数': '2710575', '销售额': '135778073.8'},
    {'标识': '每日', '日期': '5月27日', '品牌': 'Panasonic/松下', '客单价': '1619.52', '支付转化率': '1.43%', '访客数': '78928', '销售额': '1827904.286'},
    {'标识': '累计', '日期': '5月27日', '品牌': 'Panasonic/松下', '客单价': '1882.24', '支付转化率': '2.89%', '访客数': '2710575', '销售额': '135778073.8'},
    {'标识': '每日', '日期': '5月27日', '品牌': 'SIEMENS/西门子', '客单价': '1619.52', '支付转化率': '1.43%', '访客数': '78928', '销售额': '1827904.286'},
    {'标识': '累计', '日期': '5月27日', '品牌': 'SIEMENS/西门子', '客单价': '1882.24', '支付转化率': '2.89%', '访客数': '2710575', '销售额': '135778073.8'},
    {'标识': '每日', '日期': '5月27日', '品牌': 'Leader/统帅', '客单价': '1619.52', '支付转化率': '1.43%', '访客数': '78928', '销售额': '1827904.286'},
    {'标识': '累计', '日期': '5月27日', '品牌': 'Leader/统帅', '客单价': '1882.24', '支付转化率': '2.89%', '访客数': '2710575', '销售额': '135778073.8'},
    {'标识': '每日', '日期': '5月27日', '品牌': 'Royalstar/荣事达', '客单价': '1619.52', '支付转化率': '1.43%', '访客数': '78928', '销售额': '1827904.286'},
    {'标识': '累计', '日期': '5月27日', '品牌': 'Royalstar/荣事达', '客单价': '1882.24', '支付转化率': '2.89%', '访客数': '2710575', '销售额': '135778073.8'},
    {'标识': '每日', '日期': '5月27日', '品牌': 'Whirlpool/惠而浦', '客单价': '1619.52', '支付转化率': '1.43%', '访客数': '78928', '销售额': '1827904.286'},
    {'标识': '累计', '日期': '5月27日', '品牌': 'Whirlpool/惠而浦', '客单价': '1882.24', '支付转化率': '2.89%', '访客数': '2710575', '销售额': '135778073.8'},
    {'日期': '5月27日', '品牌': '美的系'},
    {'日期': '5月27日', '品牌': '惠而浦系'},
]
w = W()
for data in data_list:
    w.w(data)

'''
5月23日	美的系								
5月23日	惠而浦系	
'''

