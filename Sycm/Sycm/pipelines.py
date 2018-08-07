# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import csv
import datetime
import json
import logging
import os

import openpyxl
import xlsxwriter

from Sycm.items import SycmItem, TrendItem, SrcFlowItem
from utils.FileToDict import file_dict
from utils.format_style import xlsx_style



class SycmAllPipeline(object):
    '''数据处理：曲线走势和流量数据'''
    trend_colname = [
        '品类',
        '品牌/型号',
        '日期',
        '支付转化率',
        '支付子订单数',
        '支付件数',
    ]

    trend_col = [
        'device_category',
        'bm',
        'date_time',
        'payByrRateIndex',
        'payOrdCnt',
        'payItemQty',
    ]

    '''
    天猫搜索	淘宝搜索	直接访问	淘宝站内其他	购物车	淘宝客	聚划算	淘宝其他店铺	淘宝首页	已买到商品	其他	淘宝类目	淘外流量其他	天猫首页	我的淘宝首页
    淘内免费其他	手淘首页	手淘搜索	淘宝客	手淘品牌街	购物车	我的淘宝	手淘问大家	猫客搜索	猫客首页	手淘其他店铺商品详情	WAP天猫	手淘找相似	猫客天猫超市	聚划算	手淘消息中心	手淘拍立淘	直接访问	手淘其他店铺	手淘扫一扫	手淘我的评价	手淘微淘	直通车	手淘淘抢购
    '''

    srcflow_col = [
        '品类',
        '品牌',
        '型号',
        '商品ID',
        '天猫搜索',
        '淘宝搜索',
        '直接访问',
        '淘宝站内其他',
        '购物车',
        '淘宝客',
        '聚划算',
        '淘宝其他店铺',
        '淘宝首页',
        '已买到商品',
        '其他',
        '淘宝类目',
        '淘外流量其他',
        '天猫首页',
        '我的淘宝首页',
        '淘内免费其他',
        '手淘首页',
        '手淘搜索',
        '淘宝客',
        '手淘品牌街',
        '购物车',
        '我的淘宝',
        '手淘问大家',
        '猫客搜索',
        '猫客首页',
        '手淘其他店铺商品详情',
        'WAP天猫',
        '手淘找相似',
        '猫客天猫超市',
        '聚划算',
        '手淘消息中心',
        '手淘拍立淘',
        '直接访问',
        '手淘其他店铺',
        '手淘扫一扫',
        '手淘我的评价',
        '手淘微淘',
        '直通车',
        '手淘淘抢购',
    ]

    srcflow_colname = [
        'device_category',
        'brandName',
        'modelName',
        'itemId',
        'pc',
        'pc_uv',
        'wifi',
        'wifi_uv',
    ]

    def __init__(self):
        '''初始化文件'''
        folder = './ctmall/ctmall_%s/' % datetime.datetime.now().strftime("%Y%m%d")
        if os.path.exists(folder):
            print('exists')
        else:
            os.mkdir(folder)

        print('start write NoData')
        filename = folder + 'NoData_%s.txt' % datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        if os.path.exists(filename):
            print('remove txt start')
            os.remove(filename)
            print('remove txt end')
        # 在爬虫启动时，创建csv，并设置newline=''来避免空行出现
        print('create txt')

        self.f = open(filename, 'w')


        print('start write Trend')
        filename_trend = folder + 'Trend_%s.xlsx' % datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        if os.path.exists(filename_trend):
            print('remove xlsx start')
            os.remove(filename_trend)
            print('remove xlsx end')
        # 在爬虫启动时，创建csv，并设置newline=''来避免空行出现
        print('create xlsx')

        print('start write SrcFlow')
        filename_srcflow = folder + 'SrcFlow_%s.xlsx' % datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        if os.path.exists(filename_srcflow):
            print('remove xlsx start')
            os.remove(filename_srcflow)
            print('remove xlsx end')
        # 在爬虫启动时，创建csv，并设置newline=''来避免空行出现
        print('create xlsx')

        # print('start write SrcFlow')
        # filename_srcflow1 = folder + 'SrcFlowOld_%s.xlsx' % datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        # if os.path.exists(filename_srcflow):
        #     print('remove xlsx start')
        #     os.remove(filename_srcflow)
        #     print('remove xlsx end')
        # # 在爬虫启动时，创建csv，并设置newline=''来避免空行出现
        # print('create xlsx')

        # 创建一个新的excel文件并添加一个工作表
        self.wb_trend = xlsxwriter.Workbook(filename_trend)

        # 创建一个新的excel文件并添加一个工作表
        self.wb_srcflow = xlsxwriter.Workbook(filename_srcflow)

        # 创建一个新的excel文件并添加一个工作表
        # self.wb_srcflow1 = xlsxwriter.Workbook(filename_srcflow1)

        # 创建工作簿
        self.ws_trend = self.wb_trend.add_worksheet('曲线走势图数据')
        print('chuang jian trend')

        # 创建工作簿
        self.ws_srcflow = self.wb_srcflow.add_worksheet('流量数据改版')
        print('chuang jian srcflow')

        # 创建工作簿
        self.ws_srcflow1 = self.wb_srcflow.add_worksheet('流量数据原版')
        print('chuang jian srcflow')

        # 行，列初始值
        self.row_trend = 0
        self.col_trend = 0

        # 设置表头
        for coln in self.trend_colname:
            self.ws_trend.write(self.row_trend, self.col_trend, coln)
            self.col_trend += 1

        # 行，列初始值
        self.row_srcflow1 = 0
        self.col_srcflow1 = 0

        # 设置表头
        for coln in self.srcflow_colname:
            self.ws_srcflow1.write(self.row_srcflow1, self.col_srcflow1, coln)
            self.col_srcflow1 += 1

        # 行，列初始值
        self.row_srcflow = 1
        self.col_srcflow = 0
        # 居中格式
        center_style_pc = self.wb_srcflow.add_format(xlsx_style(bg_color='#FFFF00'))
        center_style_wifi = self.wb_srcflow.add_format(xlsx_style(bg_color='#FFCC00'))
        # 合并单元格
        self.ws_srcflow.merge_range(0, 4, 0, 18, 'pc')
        self.ws_srcflow.merge_range(0, 19, 0, 42, '无线')
        self.ws_srcflow.write(0, 4, 'pc', center_style_pc)
        self.ws_srcflow.write(0, 19, '无线', center_style_wifi)

        # 设置表头
        for coln in self.srcflow_col:
            self.ws_srcflow.write(self.row_srcflow, self.col_srcflow, coln)
            self.col_srcflow += 1

    def process_item(self, item, spider):
        '''数据处理'''
        # print('-' * 30)
        # print(item.__class__)
        if 'no' == item['info']:
        # if item.__class__ == SycmItem:

            print('NoPipeline process_item')
            json_str = json.dumps(item, ensure_ascii=False) + ',\n'
            self.f.write(json_str)

            # return item

        elif 'yes' == item['info'] and 'payByrRateIndex' in item:
        # elif item.__class__ == TrendItem:
        #     print('TrendPipeline process_item')

            # print(item['brandName'], ' ', item['modelName'])

            # 筛选
            # 整理出item字典中的所有key
            item_keys = list(item.keys())
            # 根据表头字段筛选出需要的key
            itemkey_list = [item_key for item_key in item_keys if item_key in self.trend_col]

            # 建立一个新的字典
            item_dict = dict()
            last_dict = dict()
            for itemkey in itemkey_list:
                item_dict[itemkey] = item[itemkey]

            if item['num'] == item['total'] - 1:
                item_dict['bm'] = item['brandName'] + ' ' + item['modelName']
                item_dict['device_category'] = item['device_category']

            else:
                item_dict['bm'] = ''
                item_dict['device_category'] = ''
            # print('save xlsx')

            # print(item_dict)

            last_dict['品类'] = item_dict['device_category']
            last_dict['品牌/型号'] = item_dict['bm']
            last_dict['日期'] = item_dict['date_time']
            last_dict['支付转化率'] = item_dict['payByrRateIndex']
            last_dict['支付子订单数'] = item_dict['payOrdCnt']
            last_dict['支付件数'] = item_dict['payItemQty']
            # print(last_dict)
            # 行的移动
            self.row_trend += 1
            # 写入时间数据
            # self.ws.write(self.row, 0, item_dict['date_time'])
            # 写入主要数据
            for key, value in last_dict.items():
                for index, col in enumerate(self.trend_colname):
                    if key == col:
                        self.ws_trend.write(self.row_trend, index, value)

            # return item
            # print('save TrendItem xlsx')


        elif 'yes' == item['info'] and 'i' in item:
        # elif item.__class__ == SrcFlowItem:
        #     print('SrcFlowPipeline process_item')

            pc_dict = item['pc']
            wifi_dict = item['wifi']



            max_num = max(len(pc_dict), len(wifi_dict))

            # extend合并列表
            pc_keys = list(pc_dict.keys())
            wifi_keys = list(wifi_dict.keys())

            # 筛选
            # item_keys = list(item.keys())
            # itemkey_list = [item_key for item_key in item_keys if item_key in self.colname]

            # 建立一个新的字典
            item_dict = dict()

            item_dict['品类'] = item['device_category']
            item_dict['品牌'] = item['brandName']
            item_dict['型号'] = item['modelName']
            item_dict['商品ID'] = item['itemId']

            # print('item:', item)
            # print('pc:', pc_dict, pc_keys)
            # print('wifi:', wifi_dict, wifi_keys)




            for i in range(max_num):
                item_dict1 = dict()
                if i == 0:
                    item_dict1['device_category'] = item['device_category']
                    item_dict1['brandName'] = item['brandName']
                    item_dict1['modelName'] = item['modelName']
                    item_dict1['itemId'] = item['itemId']
                else:
                    item_dict1['device_category'] = ''
                    item_dict1['brandName'] = ''
                    item_dict1['modelName'] = ''
                    item_dict1['itemId'] = ''
                try:
                    item_dict1['pc'] = pc_keys[i]
                    item_dict1['pc_uv'] = pc_dict[pc_keys[i]]
                except:
                    item_dict1['pc'] = ''
                    item_dict1['pc_uv'] = ''
                try:
                    item_dict1['wifi'] = wifi_keys[i]
                    item_dict1['wifi_uv'] = wifi_dict[wifi_keys[i]]
                except:
                    item_dict1['wifi'] = ''
                    item_dict1['wifi_uv'] = ''

                # 行的移动
                self.row_srcflow1 += 1

                # print('item_dict1:', item_dict1)

                # 写入主要数据
                for key, value in item_dict1.items():
                    for index, col in enumerate(self.srcflow_colname):
                        if key == col:
                            self.ws_srcflow1.write(self.row_srcflow1, index, value)

            # 行的移动
            self.row_srcflow += 1
            # self.row_srcflow1 += 1



            # 写入主要数据
            for key, value in item_dict.items():
                for index, col in enumerate(self.srcflow_col):
                    if key == col:
                        self.ws_srcflow.write(self.row_srcflow, index, value)
            # 写入主要数据
            for key, value in pc_dict.items():
                for index, col in enumerate(self.srcflow_col):
                    if key == col and index < 19:
                        self.ws_srcflow.write(self.row_srcflow, index, value)
            # 写入主要数据
            for key, value in wifi_dict.items():
                for index, col in enumerate(self.srcflow_col):
                    if key == col and index > 18:
                        self.ws_srcflow.write(self.row_srcflow, index, value)

            # print('save SrcFlowItem xlsx')
            # return item

    def __del__(self):
        '''关闭文件'''
        _, line_list = file_dict({'modelName': '1', 'brandName': '1'})
        self.f.write(str(line_list))
        # print(str(line_list))

        self.f.close()
        print('txt ok')

        # 在爬虫结束时，关闭文件节省资源
        logging.info('Trend finished')

        # 表格存储结束
        self.wb_trend.close()
        print('trend ok')
        # 保存数据后对表格进一步处理

        # 在爬虫结束时，关闭文件节省资源
        logging.info('SrcFlow finished')

        # 表格存储结束
        self.wb_srcflow.close()
        # self.wb_srcflow1.close()
        print('srcflow ok')

        # 保存数据后对表格进一步处理


class SanyoPipeline(object):
    colname_old = [
        '月份',
        '品牌名称',
        'Haier/海尔',
        'SIEMENS/西门子',
        'Midea/美的',
        'Samsung/三星',
        'Bosch/博世',
        'Whirlpool/惠而浦',
        'Royalstar/荣事达',
        'DIQUA/帝度',
    ]

    colname = [
        '月份',
        '品牌名称',
        'Haier/海尔',
        'Littleswan/小天鹅',
        'Midea/美的',
        'SIEMENS/西门子',
        'Sanyo/三洋',
        'Royalstar/荣事达',
        'Whirlpool/惠而浦',
        'TCL',
        'Panasonic/松下',
        'Leader/统帅',
    ]

    rowname = [
        '支付商品数',
        '支付转化率',
        '客单价',
        '访客数',
        '收藏人数',
        '加购人数',
    ]

    colname_ind = [
        '日期',
        '访客数',
        '加购人数',
        '收藏人数',
        '支付件数',
        '客单价',
        '搜索点击率',
        '搜索点击人数',
    ]


    def __init__(self):
        for i in range(1, 11):
            self.filename1 = './sanyo/02-5月洗衣机行业占比累计数据_%s.xlsx' % (datetime.date.today() - datetime.timedelta(days=i)).strftime('%Y%m%d')
            self.filename2 = './sanyo/03-5月冰箱行业占比累计数据_%s.xlsx' % (datetime.date.today() - datetime.timedelta(days=i)).strftime('%Y%m%d')
            try:
                # print(self.filename1)
                # print(self.filename2)

                self.wb_washer = openpyxl.load_workbook(self.filename1)
                self.wb_fridge = openpyxl.load_workbook(self.filename2)
            except:
                continue
            else:
                break
        # sheet_name = wb.get_sheet_names()[1]
        sheet_name_washer = self.wb_washer.sheetnames[1]
        # ws = wb.get_sheet_by_name(sheet_name)
        self.ws_washer = self.wb_washer[sheet_name_washer]
        self.ws_washer_add = self.wb_washer.create_sheet('全部原始数据')
        print(self.ws_washer)
        print(self.ws_washer['C1'].value)
        print(self.ws_washer.max_row)
        print(self.ws_washer.max_column)
        self.maxRow_washer = self.ws_washer.max_row
        self.row_washer = 1


        # sheet_name = wb.get_sheet_names()[1]
        sheet_name_fridge = self.wb_fridge.sheetnames[1]
        # ws = wb.get_sheet_by_name(sheet_name)
        self.ws_fridge = self.wb_fridge[sheet_name_fridge]
        self.ws_fridge_add = self.wb_fridge.create_sheet('全部原始数据')
        print(self.ws_fridge)
        print(self.ws_fridge['C1'].value)
        print(self.ws_fridge.max_row)
        print(self.ws_fridge.max_column)
        self.maxRow_fridge = self.ws_fridge.max_row
        self.row_fridge = 1

        # # 品牌
        # # 创建一个新的excel文件并添加一个工作表
        # self.wb = xlsxwriter.Workbook('fridge_%s.xlsx' % datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
        # # 创建工作簿
        # self.ws = self.wb.add_worksheet('洗衣机品牌数据月')
        # # print('chuang jian trend')
        #
        # self.center_style = self.wb.add_format(xlsx_style())
        #
        # # 行，列初始值
        # self.row = 0
        # self.col = 0
        #
        # # row_n = 1
        #
        # # # 设置表头洗衣机
        # # for coln in self.colname:
        # #     self.ws.write(self.row, self.col, coln, self.center_style)
        # #     self.col += 1
        #
        # # 设置表头冰箱
        # for coln in self.colname_old:
        #     self.ws.write(self.row, self.col, coln, self.center_style)
        #     self.col += 1
        #
        # self.row_h = 1
        # self.row_s = 1
        # self.row_m = 1
        # self.row_x = 1
        # self.row_b = 1
        # self.row_w = 1
        # self.row_r = 1
        # self.row_d = 1
        # # self.row_p = 1
        # # self.row_l = 1

        # # 大盘
        # # 创建一个新的excel文件并添加一个工作表
        # self.wb_ind = xlsxwriter.Workbook('dapan_%s.xlsx' % datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
        # # 创建工作簿
        # self.ws_2016 = self.wb_ind.add_worksheet('洗衣机大盘2016')
        # self.ws_2017 = self.wb_ind.add_worksheet('洗衣机大盘2017')
        # self.ws_2018 = self.wb_ind.add_worksheet('洗衣机大盘2018')
        # # print('chuang jian trend')
        #
        # self.center_style = self.wb_ind.add_format(xlsx_style())
        #
        # self.ws_2016.write(0, 0, '2016行业大盘数据指标', self.center_style)
        # self.ws_2016.merge_range(0, 0, 0, 7, '2016行业大盘数据指标')
        # self.ws_2017.write(0, 0, '2017行业大盘数据指标', self.center_style)
        # self.ws_2017.merge_range(0, 0, 0, 7, '2017行业大盘数据指标')
        # self.ws_2018.write(0, 0, '2018行业大盘数据指标', self.center_style)
        # self.ws_2018.merge_range(0, 0, 0, 7, '2018行业大盘数据指标')
        #
        # # 行，列初始值
        # self.row = 1
        # self.col = 0
        #
        # # row_n = 1
        #
        # # 设置表头
        # for coln in self.colname_ind:
        #     self.ws_2016.write(self.row, self.col, coln, self.center_style)
        #     self.ws_2017.write(self.row, self.col, coln, self.center_style)
        #     self.ws_2018.write(self.row, self.col, coln, self.center_style)
        #     self.col += 1
        #
        # self.row_2016 = 2
        # self.col_2016 = 0
        # self.row_2017 = 2
        # self.col_2017 = 0
        # self.row_2018 = 2
        # self.col_2018 = 0



    def process_item(self, item, spider):
        '''数据处理'''
        # print('-' * 30)
        # print(item.__class__)
        if 'sanyo' == spider.name:

            if 'home' == item['mark']:
                pass

            elif 'brand' == item['mark']:
                item_dict = dict()
                item_dict['日期'] = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%m月%d日')
                item_dict['品牌'] = item['brand']
                item_dict['交易指数'] = item['tradeIndex']
                item_dict['支付商品数'] = item['payItemCnt']
                item_dict['客单价'] = item['payPct']
                item_dict['支付转化率'] = item['payRate']
                item_dict['访客数'] = item['uv']
                item_dict['搜索点击人数'] = item['searchUvCnt']
                item_dict['收藏人数'] = item['favBuyerCnt']
                item_dict['加购人数'] = item['addCartUserCnt']
                item_dict['卖家数'] = item['sellerCnt']
                item_dict['被支付卖家数'] = item['paySellerCnt']
                item_dict['重点卖家数'] = item['majorSellerCnt']
                item_dict['重点商品数'] = item['majorItemCnt']
                item_dict['每日累计区分'] = item['index']
                if 'washer' == item['cate']:
                    self.row_washer += 1
                    item_list = list(item_dict.keys())
                    for item_no in range(1, len(item_list) + 1):
                        self.ws_washer_add.cell(row=1, column=item_no, value=item_list[item_no - 1])
                        self.ws_washer_add.cell(row=self.row_washer, column=item_no, value=item_dict[item_list[item_no - 1]])
                    data = dict()
                    data['标识'] = item['index']
                    data['日期'] = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%m月%d日')
                    data['品牌'] = item['brand']
                    data['客单价'] = item['payPct']
                    data['支付转化率'] = str(round(item['payRate'] * 100, 2)) + '%'
                    data['访客数'] = item['uv']
                    data['销售额'] = item['payPct'] * item['payRate'] * item['uv']
                    print(data)
                    data_keys = list(data.keys())
                    print(data_keys)
                    if '1' == data['标识']:
                        if 'Haier/海尔' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 2, column=i, value=data[data_keys[i]])
                        elif 'Littleswan/小天鹅' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 5, column=i, value=data[data_keys[i]])
                        elif 'Midea/美的' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 6, column=i, value=data[data_keys[i]])
                        elif 'Sanyo/三洋' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 7, column=i, value=data[data_keys[i]])
                        elif 'TCL' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 8, column=i, value=data[data_keys[i]])
                        elif 'Panasonic/松下' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 9, column=i, value=data[data_keys[i]])
                        elif 'SIEMENS/西门子' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 10, column=i, value=data[data_keys[i]])
                        elif 'Leader/统帅' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 11, column=i, value=data[data_keys[i]])
                        elif 'Royalstar/荣事达' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 12, column=i, value=data[data_keys[i]])
                        elif 'Whirlpool/惠而浦' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 13, column=i, value=data[data_keys[i]])

                    elif '4' == data['标识']:
                        if 'Haier/海尔' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 2, column=i, value=data[data_keys[i - 4]])
                        elif 'Littleswan/小天鹅' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 5, column=i, value=data[data_keys[i - 4]])
                        elif 'Midea/美的' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 6, column=i, value=data[data_keys[i - 4]])
                        elif 'Sanyo/三洋' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 7, column=i, value=data[data_keys[i - 4]])
                        elif 'TCL' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 8, column=i, value=data[data_keys[i - 4]])
                        elif 'Panasonic/松下' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 9, column=i, value=data[data_keys[i - 4]])
                        elif 'SIEMENS/西门子' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 10, column=i, value=data[data_keys[i - 4]])
                        elif 'Leader/统帅' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 11, column=i, value=data[data_keys[i - 4]])
                        elif 'Royalstar/荣事达' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 12, column=i, value=data[data_keys[i - 4]])
                        elif 'Whirlpool/惠而浦' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 13, column=i, value=data[data_keys[i - 4]])



                    self.ws_washer.cell(row=self.maxRow_washer + 3, column=1, value=data['日期'])
                    self.ws_washer.cell(row=self.maxRow_washer + 3, column=2, value='美的系')

                    self.ws_washer.cell(row=self.maxRow_washer + 4, column=1, value=data['日期'])
                    self.ws_washer.cell(row=self.maxRow_washer + 4, column=2, value='惠而浦系')

                    pass
                elif 'fridge' == item['cate']:
                    self.row_fridge += 1
                    item_list = list(item_dict.keys())
                    for item_no in range(1, len(item_list) + 1):
                        self.ws_fridge_add.cell(row=1, column=item_no, value=item_list[item_no - 1])
                        self.ws_fridge_add.cell(row=self.row_fridge, column=item_no, value=item_dict[item_list[item_no -1]])
                    data = dict()
                    data['标识'] = item['index']
                    data['日期'] = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%m月%d日')
                    data['品牌'] = item['brand']
                    data['客单价'] = item['payPct']
                    data['支付转化率'] = str(round(item['payRate'] * 100, 2)) + '%'
                    data['访客数'] = item['uv']
                    data['销售额'] = item['payPct'] * item['payRate'] * item['uv']
                    print(data)
                    data_keys = list(data.keys())
                    print(data_keys)

                    if '标识' in data and '1' == data['标识']:
                        if 'Haier/海尔' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 2, column=i, value=data[data_keys[i]])
                        elif 'SIEMENS/西门子' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 3, column=i, value=data[data_keys[i]])
                        elif 'Midea/美的' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 4, column=i, value=data[data_keys[i]])
                        elif 'Samsung/三星' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 5, column=i, value=data[data_keys[i]])
                        elif 'Bosch/博世' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 6, column=i, value=data[data_keys[i]])
                        elif 'Whirlpool/惠而浦' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 7, column=i, value=data[data_keys[i]])
                        elif 'Royalstar/荣事达' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 8, column=i, value=data[data_keys[i]])
                        elif 'DIQUA/帝度' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 9, column=i, value=data[data_keys[i]])

                    elif '标识' in data and '4' == data['标识']:
                        if 'Haier/海尔' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 2, column=i, value=data[data_keys[i - 4]])
                        elif 'SIEMENS/西门子' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 3, column=i, value=data[data_keys[i - 4]])
                        elif 'Midea/美的' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 4, column=i, value=data[data_keys[i - 4]])
                        elif 'Samsung/三星' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 5, column=i, value=data[data_keys[i - 4]])
                        elif 'Bosch/博世' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 6, column=i, value=data[data_keys[i - 4]])
                        elif 'Whirlpool/惠而浦' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 7, column=i, value=data[data_keys[i - 4]])
                        elif 'Royalstar/荣事达' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 8, column=i, value=data[data_keys[i - 4]])
                        elif 'DIQUA/帝度' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 9, column=i, value=data[data_keys[i - 4]])



                # if 'fridge' == item['cate']:
                #     # 品牌
                #     item_dict = dict()
                #     item_dict['支付商品数'] = item['payItemCnt']
                #     item_dict['支付转化率'] = str(round(item['payRate'] * 100, 2)) + '%'
                #     item_dict['客单价'] = item['payPct']
                #     item_dict['访客数'] = item['uv']
                #     item_dict['收藏人数'] = item['favBuyerCnt']
                #     item_dict['加购人数'] = item['addCartUserCnt']
                #
                #     # print(item_dict)
                #     # Haier/海尔	SIEMENS/西门子	Midea/美的	Samsung/三星	Bosch/博世	Whirlpool/惠而浦	Royalstar/荣事达	DIQUA/帝度
                #     # 整理出item字典中的所有key
                #     item_keys = list(item_dict.keys())
                #     # print(item_keys)
                #
                #     year_month = '%s年%s月' % (item['year'], item['month'])
                #
                #     # 冰箱
                #     if 'Haier/海尔' == item['brand']:
                #
                #         self.ws.merge_range(self.row_h, 0, self.row_h + 5, 0, year_month, self.center_style)
                #         print(item['brand'], year_month)
                #
                #         self.ws.write(self.row_h, 0, year_month, self.center_style)
                #         for rown in self.rowname:
                #             self.ws.write(self.row_h, 1, rown, self.center_style)
                #             # for item_key in item_keys:
                #             #     if item_key == rown:
                #             self.ws.write(self.row_h, 2, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #             self.row_h += 1
                #             # print(self.row_h)
                #
                #         pass
                #     elif 'SIEMENS/西门子' == item['brand']:
                #         # self.ws.merge_range(self.row_s, 0, self.row_s + 5, 0, year_month, self.center_style)
                #         print(item['brand'], year_month)
                #         # self.ws.write(self.row_s, 0, year_month, self.center_style)
                #         for rown in self.rowname:
                #             # self.ws.write(self.row_s, 1, rown, self.center_style)
                #             # for item_key in item_keys:
                #             #     if item_key == rown:
                #             self.ws.write(self.row_s, 3, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #             self.row_s += 1
                #             # print(self.row_s)
                #         pass
                #     elif 'Midea/美的' == item['brand']:
                #         # self.ws.merge_range(self.row_m, 0, self.row_m + 5, 0, year_month, self.center_style)
                #         print(item['brand'], year_month)
                #         # self.ws.write(self.row_m, 0, year_month, self.center_style)
                #         for rown in self.rowname:
                #             # self.ws.write(self.row_m, 1, rown, self.center_style)
                #             # for item_key in item_keys:
                #             #     if item_key == rown:
                #             self.ws.write(self.row_m, 4, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #             self.row_m += 1
                #             # print(self.row_m)
                #         pass
                #     elif 'Samsung/三星' == item['brand']:
                #         # self.ws.merge_range(self.row_x, 0, self.row_x + 5, 0, year_month, self.center_style)
                #         print(item['brand'], year_month)
                #         # self.ws.write(self.row_x, 0, year_month, self.center_style)
                #         for rown in self.rowname:
                #             # self.ws.write(self.row_x, 1, rown, self.center_style)
                #             # for item_key in item_keys:
                #             #     if item_key == rown:
                #             self.ws.write(self.row_x, 5, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #             self.row_x += 1
                #             # print(self.row_x)
                #         pass
                #     elif 'Bosch/博世' == item['brand']:
                #         # self.ws.merge_range(self.row_b, 0, self.row_b + 5, 0, year_month, self.center_style)
                #         print(item['brand'], year_month)
                #         # self.ws.write(self.row_b, 0, year_month, self.center_style)
                #         for rown in self.rowname:
                #             # self.ws.write(self.row_b, 1, rown, self.center_style)
                #             # for item_key in item_keys:
                #             #     if item_key == rown:
                #             self.ws.write(self.row_b, 6, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #             self.row_b += 1
                #             # print(self.row_b)
                #         pass
                #     elif 'Whirlpool/惠而浦' == item['brand']:
                #         # self.ws.merge_range(self.row_w, 0, self.row_w + 5, 0, year_month, self.center_style)
                #         print(item['brand'], year_month)
                #         # self.ws.write(self.row_w, 0, year_month, self.center_style)
                #         for rown in self.rowname:
                #             # self.ws.write(self.row_w, 1, rown, self.center_style)
                #             # for item_key in item_keys:
                #             #     if item_key == rown:
                #             self.ws.write(self.row_w, 7, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #             self.row_w += 1
                #             # print(self.row_w)
                #         pass
                #     elif 'Royalstar/荣事达' == item['brand']:
                #         # self.ws.merge_range(self.row_r, 0, self.row_r + 5, 0, year_month, self.center_style)
                #         print(item['brand'], year_month)
                #         # self.ws.write(self.row_r, 0, year_month, self.center_style)
                #         for rown in self.rowname:
                #             # self.ws.write(self.row_r, 1, rown, self.center_style)
                #             # for item_key in item_keys:
                #             #     if item_key == rown:
                #             self.ws.write(self.row_r, 8, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #             self.row_r += 1
                #             # print(self.row_r)
                #         pass
                #     elif 'DIQUA/帝度' == item['brand']:
                #         # self.ws.merge_range(self.row_d, 0, self.row_d + 5, 0, year_month, self.center_style)
                #         print(item['brand'], year_month)
                #         # self.ws.write(self.row_d, 0, year_month, self.center_style)
                #         for rown in self.rowname:
                #             # self.ws.write(self.row_d, 1, rown, self.center_style)
                #             # for item_key in item_keys:
                #             #     if item_key == rown:
                #             self.ws.write(self.row_d, 9, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #             self.row_d += 1
                #             # print(self.row_d)
                #         pass
                #
                #
                #     # # 洗衣机
                #     # if 'Haier/海尔' == item['brand']:
                #     #
                #     #     self.ws.merge_range(self.row_h, 0, self.row_h + 5, 0, year_month, self.center_style)
                #     #     print(item['brand'], year_month)
                #     #
                #     #     self.ws.write(self.row_h, 0, year_month, self.center_style)
                #     #     for rown in self.rowname:
                #     #         self.ws.write(self.row_h, 1, rown, self.center_style)
                #     #         # for item_key in item_keys:
                #     #         #     if item_key == rown:
                #     #         self.ws.write(self.row_h, 2, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #     #         self.row_h += 1
                #     #         # print(self.row_h)
                #     #
                #     #     pass
                #     # elif 'Littleswan/小天鹅' == item['brand']:
                #     #     # self.ws.merge_range(self.row_s, 0, self.row_s + 5, 0, year_month, self.center_style)
                #     #     print(item['brand'], year_month)
                #     #     # self.ws.write(self.row_s, 0, year_month, self.center_style)
                #     #     for rown in self.rowname:
                #     #         # self.ws.write(self.row_s, 1, rown, self.center_style)
                #     #         # for item_key in item_keys:
                #     #         #     if item_key == rown:
                #     #         self.ws.write(self.row_s, 3, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #     #         self.row_s += 1
                #     #         # print(self.row_s)
                #     #     pass
                #     # elif 'Midea/美的' == item['brand']:
                #     #     # self.ws.merge_range(self.row_m, 0, self.row_m + 5, 0, year_month, self.center_style)
                #     #     print(item['brand'], year_month)
                #     #     # self.ws.write(self.row_m, 0, year_month, self.center_style)
                #     #     for rown in self.rowname:
                #     #         # self.ws.write(self.row_m, 1, rown, self.center_style)
                #     #         # for item_key in item_keys:
                #     #         #     if item_key == rown:
                #     #         self.ws.write(self.row_m, 4, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #     #         self.row_m += 1
                #     #         # print(self.row_m)
                #     #     pass
                #     # elif 'SIEMENS/西门子' == item['brand']:
                #     #     # self.ws.merge_range(self.row_x, 0, self.row_x + 5, 0, year_month, self.center_style)
                #     #     print(item['brand'], year_month)
                #     #     # self.ws.write(self.row_x, 0, year_month, self.center_style)
                #     #     for rown in self.rowname:
                #     #         # self.ws.write(self.row_x, 1, rown, self.center_style)
                #     #         # for item_key in item_keys:
                #     #         #     if item_key == rown:
                #     #         self.ws.write(self.row_x, 5, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #     #         self.row_x += 1
                #     #         # print(self.row_x)
                #     #     pass
                #     # elif 'Sanyo/三洋' == item['brand']:
                #     #     # self.ws.merge_range(self.row_b, 0, self.row_b + 5, 0, year_month, self.center_style)
                #     #     print(item['brand'], year_month)
                #     #     # self.ws.write(self.row_b, 0, year_month, self.center_style)
                #     #     for rown in self.rowname:
                #     #         # self.ws.write(self.row_b, 1, rown, self.center_style)
                #     #         # for item_key in item_keys:
                #     #         #     if item_key == rown:
                #     #         self.ws.write(self.row_b, 6, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #     #         self.row_b += 1
                #     #         # print(self.row_b)
                #     #     pass
                #     # elif 'Royalstar/荣事达' == item['brand']:
                #     #     # self.ws.merge_range(self.row_w, 0, self.row_w + 5, 0, year_month, self.center_style)
                #     #     print(item['brand'], year_month)
                #     #     # self.ws.write(self.row_w, 0, year_month, self.center_style)
                #     #     for rown in self.rowname:
                #     #         # self.ws.write(self.row_w, 1, rown, self.center_style)
                #     #         # for item_key in item_keys:
                #     #         #     if item_key == rown:
                #     #         self.ws.write(self.row_w, 7, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #     #         self.row_w += 1
                #     #         # print(self.row_w)
                #     #     pass
                #     # elif 'Whirlpool/惠而浦' == item['brand']:
                #     #     # self.ws.merge_range(self.row_r, 0, self.row_r + 5, 0, year_month, self.center_style)
                #     #     print(item['brand'], year_month)
                #     #     # self.ws.write(self.row_r, 0, year_month, self.center_style)
                #     #     for rown in self.rowname:
                #     #         # self.ws.write(self.row_r, 1, rown, self.center_style)
                #     #         # for item_key in item_keys:
                #     #         #     if item_key == rown:
                #     #         self.ws.write(self.row_r, 8, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #     #         self.row_r += 1
                #     #         # print(self.row_r)
                #     #     pass
                #     # elif 'TCL' == item['brand']:
                #     #     # self.ws.merge_range(self.row_d, 0, self.row_d + 5, 0, year_month, self.center_style)
                #     #     print(item['brand'], year_month)
                #     #     # self.ws.write(self.row_d, 0, year_month, self.center_style)
                #     #     for rown in self.rowname:
                #     #         # self.ws.write(self.row_d, 1, rown, self.center_style)
                #     #         # for item_key in item_keys:
                #     #         #     if item_key == rown:
                #     #         self.ws.write(self.row_d, 9, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #     #         self.row_d += 1
                #     #         # print(self.row_d)
                #     #     pass
                #     # elif 'Panasonic/松下' == item['brand']:
                #     #     # self.ws.merge_range(self.row_d, 0, self.row_d + 5, 0, year_month, self.center_style)
                #     #     print(item['brand'], year_month)
                #     #     # self.ws.write(self.row_d, 0, year_month, self.center_style)
                #     #     for rown in self.rowname:
                #     #         # self.ws.write(self.row_d, 1, rown, self.center_style)
                #     #         # for item_key in item_keys:
                #     #         #     if item_key == rown:
                #     #         self.ws.write(self.row_p, 10, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #     #         self.row_p += 1
                #     #         # print(self.row_d)
                #     #     pass
                #     # elif 'Leader/统帅' == item['brand']:
                #     #     # self.ws.merge_range(self.row_d, 0, self.row_d + 5, 0, year_month, self.center_style)
                #     #     print(item['brand'], year_month)
                #     #     # self.ws.write(self.row_d, 0, year_month, self.center_style)
                #     #     for rown in self.rowname:
                #     #         # self.ws.write(self.row_d, 1, rown, self.center_style)
                #     #         # for item_key in item_keys:
                #     #         #     if item_key == rown:
                #     #         self.ws.write(self.row_l, 11, item_dict[item_keys[item_keys.index(rown)]], self.center_style)
                #     #         self.row_l += 1
                #     #         # print(self.row_d)
                #     #     pass
                pass

            elif 'industry' == item['mark']:
                if 'washer' == item['cate']:
                    # 日期	访客数	加购人数	收藏人数	支付件数	客单价	搜索点击率	搜索点击人数
                    item_dict = dict()
                    item_dict['访客数'] = item['uv']
                    item_dict['加购人数'] = item['addCartBuyerCnt']
                    item_dict['收藏人数'] = item['favBuyerCnt']
                    item_dict['支付件数'] = item['payItemQty']
                    item_dict['客单价'] = item['payPct']
                    item_dict['搜索点击率'] = item['searchClkRate']
                    item_dict['搜索点击人数'] = item['searchUvCnt']

                    # 整理出item字典中的所有key
                    item_keys = list(item_dict.keys())

                    year_month_day = '%s年%s月%s日' % (item['year'], item['month'], item['day'])

                    if 2016 == item['year']:
                        print('2016ri qi:', year_month_day)

                        self.ws_2016.write(self.row_2016, self.col_2016, year_month_day, self.center_style)
                        for item_key in item_keys:

                            self.ws_2016.write(self.row_2016, self.colname_ind.index(item_key), item_dict[item_key], self.center_style)

                        self.row_2016 += 1

                        pass
                    elif 2017 == item['year']:
                        print('2017ri qi:', year_month_day)

                        self.ws_2017.write(self.row_2017, self.col_2017, year_month_day, self.center_style)
                        for item_key in item_keys:
                            self.ws_2017.write(self.row_2017, self.colname_ind.index(item_key), item_dict[item_key],
                                               self.center_style)

                        self.row_2017 += 1

                        pass
                    elif 2018 == item['year']:
                        print('2018ri qi:', year_month_day)

                        self.ws_2018.write(self.row_2018, self.col_2018, year_month_day, self.center_style)
                        for item_key in item_keys:
                            self.ws_2018.write(self.row_2018, self.colname_ind.index(item_key), item_dict[item_key],
                                               self.center_style)

                        self.row_2018 += 1
                        pass


                    pass
                elif 'fridge' == item['cate']:
                    pass
                pass




        # elif 'sanyoRT' == spider.name:
        # # elif item.__class__ == TrendItem:
        # #     print('TrendPipeline process_item')
        #
        #     # print(item['brandName'], ' ', item['modelName'])
        #
        #     # 筛选
        #     # 整理出item字典中的所有key
        #     item_keys = list(item.keys())
        #     # 根据表头字段筛选出需要的key
        #     itemkey_list = [item_key for item_key in item_keys if item_key in self.trend_col]
        #
        #     # 建立一个新的字典
        #     item_dict = dict()
        #     last_dict = dict()
        #     for itemkey in itemkey_list:
        #         item_dict[itemkey] = item[itemkey]
        #
        #     if item['num'] == item['total'] - 1:
        #         item_dict['bm'] = item['brandName'] + ' ' + item['modelName']
        #         item_dict['device_category'] = item['device_category']
        #
        #     else:
        #         item_dict['bm'] = ''
        #         item_dict['device_category'] = ''
        #     # print('save xlsx')
        #
        #     # print(item_dict)
        #
        #     last_dict['品类'] = item_dict['device_category']
        #     last_dict['品牌/型号'] = item_dict['bm']
        #     last_dict['日期'] = item_dict['date_time']
        #     last_dict['支付转化率'] = item_dict['payByrRateIndex']
        #     last_dict['支付子订单数'] = item_dict['payOrdCnt']
        #     last_dict['支付件数'] = item_dict['payItemQty']
        #     # print(last_dict)
        #     # 行的移动
        #     self.row_trend += 1
        #     # 写入时间数据
        #     # self.ws.write(self.row, 0, item_dict['date_time'])
        #     # 写入主要数据
        #     for key, value in last_dict.items():
        #         for index, col in enumerate(self.trend_colname):
        #             if key == col:
        #                 self.ws_trend.write(self.row_trend, index, value)
        #
        #     # return item
        #     # print('save TrendItem xlsx')
        #
        #
        # elif 'yes' == item['info'] and 'i' in item:
        # # elif item.__class__ == SrcFlowItem:
        # #     print('SrcFlowPipeline process_item')
        #
        #     pc_dict = item['pc']
        #     wifi_dict = item['wifi']
        #
        #
        #
        #     max_num = max(len(pc_dict), len(wifi_dict))
        #
        #     # extend合并列表
        #     pc_keys = list(pc_dict.keys())
        #     wifi_keys = list(wifi_dict.keys())
        #
        #     # 筛选
        #     # item_keys = list(item.keys())
        #     # itemkey_list = [item_key for item_key in item_keys if item_key in self.colname]
        #
        #     # 建立一个新的字典
        #     item_dict = dict()
        #
        #     item_dict['品类'] = item['device_category']
        #     item_dict['品牌'] = item['brandName']
        #     item_dict['型号'] = item['modelName']
        #     item_dict['商品ID'] = item['itemId']
        #
        #     # print('item:', item)
        #     # print('pc:', pc_dict, pc_keys)
        #     # print('wifi:', wifi_dict, wifi_keys)
        #
        #
        #
        #
        #     for i in range(max_num):
        #         item_dict1 = dict()
        #         if i == 0:
        #             item_dict1['device_category'] = item['device_category']
        #             item_dict1['brandName'] = item['brandName']
        #             item_dict1['modelName'] = item['modelName']
        #             item_dict1['itemId'] = item['itemId']
        #         else:
        #             item_dict1['device_category'] = ''
        #             item_dict1['brandName'] = ''
        #             item_dict1['modelName'] = ''
        #             item_dict1['itemId'] = ''
        #         try:
        #             item_dict1['pc'] = pc_keys[i]
        #             item_dict1['pc_uv'] = pc_dict[pc_keys[i]]
        #         except:
        #             item_dict1['pc'] = ''
        #             item_dict1['pc_uv'] = ''
        #         try:
        #             item_dict1['wifi'] = wifi_keys[i]
        #             item_dict1['wifi_uv'] = wifi_dict[wifi_keys[i]]
        #         except:
        #             item_dict1['wifi'] = ''
        #             item_dict1['wifi_uv'] = ''
        #
        #         # 行的移动
        #         self.row_srcflow1 += 1
        #
        #         # print('item_dict1:', item_dict1)
        #
        #         # 写入主要数据
        #         for key, value in item_dict1.items():
        #             for index, col in enumerate(self.srcflow_colname):
        #                 if key == col:
        #                     self.ws_srcflow1.write(self.row_srcflow1, index, value)
        #
        #     # 行的移动
        #     self.row_srcflow += 1
        #     # self.row_srcflow1 += 1
        #
        #
        #
        #     # 写入主要数据
        #     for key, value in item_dict.items():
        #         for index, col in enumerate(self.srcflow_col):
        #             if key == col:
        #                 self.ws_srcflow.write(self.row_srcflow, index, value)
        #     # 写入主要数据
        #     for key, value in pc_dict.items():
        #         for index, col in enumerate(self.srcflow_col):
        #             if key == col and index < 19:
        #                 self.ws_srcflow.write(self.row_srcflow, index, value)
        #     # 写入主要数据
        #     for key, value in wifi_dict.items():
        #         for index, col in enumerate(self.srcflow_col):
        #             if key == col and index > 18:
        #                 self.ws_srcflow.write(self.row_srcflow, index, value)
        #
        #     # print('save SrcFlowItem xlsx')
        #     # return item


    # pass
    def __del__(self):
        # self.wb.close()
        # self.wb_ind.close()
        self.wb_washer.save('./sanyo/02-5月洗衣机行业占比累计数据_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d'))
        self.wb_fridge.save('./sanyo/03-5月冰箱行业占比累计数据_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d'))
        #
        os.remove(self.filename1)
        os.remove(self.filename2)


class MideaPipeline(object):

    def __init__(self):
        '''初始化'''
        if os.path.exists('./midea'):
            pass
        else:
            os.mkdir('./midea')
        self.filename = './midea/竞品日数据跟踪_%s.xlsx' % (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m')
        # self.filename = './midea/竞品日数据跟踪_20180408_20180630.xlsx'
        if os.path.exists(self.filename):
            self.wb = openpyxl.load_workbook(self.filename)

        else:
            self.wb = openpyxl.Workbook()
        # 炒锅
        # try:
        #     self.ws_cg = self.wb['炒锅']
        # 煎锅
        # 奶锅
        # 汤锅
        # 蒸锅
        try:
            self.ws_cg = self.wb['炒锅']
            self.ws_jg = self.wb['煎锅']
            self.ws_ng = self.wb['奶锅']
            self.ws_tg = self.wb['汤锅']
            self.ws_zg = self.wb['蒸锅']
        except:
            self.ws_cg = self.wb.active
            self.ws_jg = self.wb.create_sheet()
            self.ws_ng = self.wb.create_sheet()
            self.ws_tg = self.wb.create_sheet()
            self.ws_zg = self.wb.create_sheet()
            # self.ws = self.wb.active
            self.ws_cg.title = '炒锅'
            self.ws_jg.title = '煎锅'
            self.ws_ng.title = '奶锅'
            self.ws_tg.title = '汤锅'
            self.ws_zg.title = '蒸锅'

            self.ws_cg = self.wb['炒锅']
            self.ws_jg = self.wb['煎锅']
            self.ws_ng = self.wb['奶锅']
            self.ws_tg = self.wb['汤锅']
            self.ws_zg = self.wb['蒸锅']

            for ws in [self.ws_cg, self.ws_jg, self.ws_ng, self.ws_tg, self.ws_zg]:
                ws.cell(row=1, column=1, value='日期1')
                ws.cell(row=1, column=2, value='行业-访客数')
                ws.cell(row=1, column=3, value='行业-支付件数')
                ws.cell(row=1, column=4, value='行业-客单价')
                ws.cell(row=1, column=5, value='行业-搜索点击人数')
                ws.cell(row=1, column=6, value='日期2')
                ws.cell(row=1, column=7, value='美的-访客数')
                # ws.cell(row=1, column=8, value='美的-支付件数')
                ws.cell(row=1, column=8, value='美的-客单价')
                ws.cell(row=1, column=9, value='美的-支付转化率')
                ws.cell(row=1, column=10, value='美的-搜索点击人数')
                ws.cell(row=1, column=11, value='美的-支付件数')
                ws.cell(row=1, column=12, value='日期2')
                ws.cell(row=1, column=13, value='苏泊尔-访客数')

                ws.cell(row=1, column=14, value='苏泊尔-客单价')
                ws.cell(row=1, column=15, value='苏泊尔-支付转化率')
                ws.cell(row=1, column=16, value='苏泊尔-搜索点击人数')
                ws.cell(row=1, column=17, value='苏泊尔-支付件数')
                ws.cell(row=1, column=18, value='日期2')
                ws.cell(row=1, column=19, value='炊大皇-访客数')

                ws.cell(row=1, column=20, value='炊大皇-客单价')
                ws.cell(row=1, column=21, value='炊大皇-支付转化率')
                ws.cell(row=1, column=22, value='炊大皇-搜索点击人数')
                ws.cell(row=1, column=23, value='炊大皇-支付件数')
                ws.cell(row=1, column=24, value='日期2')
                ws.cell(row=1, column=25, value='爱仕达-访客数')

                ws.cell(row=1, column=26, value='爱仕达-客单价')
                ws.cell(row=1, column=27, value='爱仕达-支付转化率')
                ws.cell(row=1, column=28, value='爱仕达-搜索点击人数')
                ws.cell(row=1, column=29, value='爱仕达-支付件数')
                if self.ws_cg == ws:
                    ws.cell(row=1, column=30, value='日期2')
                    ws.cell(row=1, column=31, value='臻三环-访客数')

                    ws.cell(row=1, column=32, value='臻三环-客单价')
                    ws.cell(row=1, column=33, value='臻三环-支付转化率')
                    ws.cell(row=1, column=34, value='臻三环-搜索点击人数')
                    ws.cell(row=1, column=35, value='臻三环-支付件数')
                    ws.cell(row=1, column=36, value='日期2')
                    ws.cell(row=1, column=37, value='feillers-访客数')

                    ws.cell(row=1, column=28, value='feillers-客单价')
                    ws.cell(row=1, column=39, value='feillers-支付转化率')
                    ws.cell(row=1, column=40, value='feillers-搜索点击人数')
                    ws.cell(row=1, column=41, value='feillers-支付件数')

        self.num_cg = 0
        self.num_jg = 0
        self.num_ng = 0
        self.num_tg = 0
        self.num_zg = 0

    def process_item(self, item, spider):
        '''数据存储分类'''
        print('pipeline==========', item['time_date'])
        if '炒锅' == item['category']:

            self.save_item(self.ws_cg, item, self.num_cg)
            self.num_cg += 1
        elif '煎锅' == item['category']:

            self.save_item(self.ws_jg, item, self.num_jg)
            self.num_jg += 1
        elif '奶锅' == item['category']:

            self.save_item(self.ws_ng, item, self.num_ng)
            self.num_ng += 1
        elif '汤锅' == item['category']:

            self.save_item(self.ws_tg, item, self.num_tg)
            self.num_tg += 1
        elif '蒸锅' == item['category']:

            self.save_item(self.ws_zg, item, self.num_zg)
            self.num_zg += 1

    def save_item(self, ws, data, num):
        '''数据存储'''
        if data:
            print('save=========', data['category'])
            if num:
                maxRow = ws.max_row

            else:
                maxRow = ws.max_row + 1
            # maxCol = ws.max_column
            # data_day = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m%d')

            # 将data['time_date'] = '2018-01-01'转换成data_day = '2018/1/1'格式
            final_date = [str(int(no)) for no in data['time_date'].split('-')]
            data_day = '/'.join(final_date)

            if 'dapan' == data['mark']:

                ws.cell(row=maxRow, column=1, value=data_day)
                ws.cell(row=maxRow, column=2, value=data['uv'])
                ws.cell(row=maxRow, column=3, value=data['payItemQty'])
                ws.cell(row=maxRow, column=4, value=data['payPct'])
                ws.cell(row=maxRow, column=5, value=data['searchUvCnt'])

            else:

                ws.cell(row=maxRow, column=6, value=data_day)
                ws.cell(row=maxRow, column=7, value=data[30652]['uv'])

                ws.cell(row=maxRow, column=8, value=data[30652]['payPct'])
                ws.cell(row=maxRow, column=9, value=data[30652]['payRate'])
                ws.cell(row=maxRow, column=10, value=data[30652]['searchUvCnt'])
                ws.cell(row=maxRow, column=11, value=data[30652]['payItemQty'])
                ws.cell(row=maxRow, column=12, value=data_day)
                ws.cell(row=maxRow, column=13, value=data[30844]['uv'])

                ws.cell(row=maxRow, column=14, value=data[30844]['payPct'])
                ws.cell(row=maxRow, column=15, value=data[30844]['payRate'])
                ws.cell(row=maxRow, column=16, value=data[30844]['searchUvCnt'])
                ws.cell(row=maxRow, column=17, value=data[30844]['payItemQty'])
                ws.cell(row=maxRow, column=18, value=data_day)
                ws.cell(row=maxRow, column=19, value=data[5725958]['uv'])

                ws.cell(row=maxRow, column=20, value=data[5725958]['payPct'])
                ws.cell(row=maxRow, column=21, value=data[5725958]['payRate'])
                ws.cell(row=maxRow, column=22, value=data[5725958]['searchUvCnt'])
                ws.cell(row=maxRow, column=23, value=data[5725958]['payItemQty'])
                ws.cell(row=maxRow, column=24, value=data_day)
                ws.cell(row=maxRow, column=25, value=data[3222885]['uv'])

                ws.cell(row=maxRow, column=26, value=data[3222885]['payPct'])
                ws.cell(row=maxRow, column=27, value=data[3222885]['payRate'])
                ws.cell(row=maxRow, column=28, value=data[3222885]['searchUvCnt'])
                ws.cell(row=maxRow, column=29, value=data[3222885]['payItemQty'])
                if '炒锅' == data['category']:
                    ws.cell(row=maxRow, column=30, value=data_day)
                    ws.cell(row=maxRow, column=31, value=data[531544226]['uv'])

                    ws.cell(row=maxRow, column=32, value=data[531544226]['payPct'])
                    ws.cell(row=maxRow, column=33, value=data[531544226]['payRate'])
                    ws.cell(row=maxRow, column=34, value=data[531544226]['searchUvCnt'])
                    ws.cell(row=maxRow, column=35, value=data[531544226]['payItemQty'])
                    ws.cell(row=maxRow, column=36, value=data_day)
                    ws.cell(row=maxRow, column=37, value=data[1344359932]['uv'])

                    ws.cell(row=maxRow, column=38, value=data[1344359932]['payPct'])
                    ws.cell(row=maxRow, column=39, value=data[1344359932]['payRate'])
                    ws.cell(row=maxRow, column=40, value=data[1344359932]['searchUvCnt'])
                    ws.cell(row=maxRow, column=41, value=data[1344359932]['payItemQty'])

            self.wb.save(self.filename)


